"""
find_career_urls.py

Reads company names from data/companies.xlsx
(Current version only prints company names.)
"""

from openpyxl import load_workbook

# Load Excel
wb = load_workbook("data/companies.xlsx")
ws = wb["Companies"]

# Print all company names
for row in range(2, ws.max_row + 1):
    company = ws[f"A{row}"].value

    if company:
        print(company)

# Save workbook
wb.save("data/companies.xlsx")