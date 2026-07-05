"""
Master Job Database
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook


class JobDatabase:

    FILE = "data/jobs_database.xlsx"

    HEADERS = [
        "Job ID",
        "Company",
        "Title",
        "Country",
        "Location",
        "ATS",
        "URL",
        "Description",
        "First Seen",
        "Last Seen",
        "Days Active",
        "Status"
    ]

    def __init__(self):

        if not os.path.exists(self.FILE):

            wb = Workbook()

            ws = wb.active

            ws.title = "Jobs"

            ws.append(self.HEADERS)

            wb.save(self.FILE)

    def clean(self, value):

        if value is None:
            return ""

        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(v) for v in value)

        if isinstance(value, dict):
            return str(value)

        return str(value)

    def update_database(self, jobs):

        wb = load_workbook(self.FILE)

        ws = wb.active

        today = datetime.today().strftime("%Y-%m-%d")

        existing = {}

        for row in range(2, ws.max_row + 1):

            key = (
                self.clean(ws.cell(row, 2).value),
                self.clean(ws.cell(row, 3).value),
                self.clean(ws.cell(row, 7).value)
            )

            existing[key] = row

        for job in jobs:

            company = self.clean(job.company)
            title = self.clean(job.title)
            country = self.clean(job.country)
            location = self.clean(job.location)
            ats = self.clean(job.ats)
            url = self.clean(job.url)
            description = self.clean(job.description)

            key = (
                company,
                title,
                url
            )

            if key in existing:

                r = existing[key]

                ws.cell(r, 10).value = today
                ws.cell(r, 12).value = "Active"

            else:

                ws.append([

                    ws.max_row,

                    company,

                    title,

                    country,

                    location,

                    ats,

                    url,

                    description,

                    today,

                    today,

                    1,

                    "New"

                ])

        for row in range(2, ws.max_row + 1):

            try:

                first = datetime.strptime(
                    str(ws.cell(row, 9).value),
                    "%Y-%m-%d"
                )

                last = datetime.strptime(
                    str(ws.cell(row, 10).value),
                    "%Y-%m-%d"
                )

                ws.cell(row, 11).value = (last - first).days + 1

            except:
                pass

        wb.save(self.FILE)

        print(f"Database Updated : {len(jobs)} jobs")