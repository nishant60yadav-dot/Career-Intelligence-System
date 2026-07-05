"""
AI Job Ranker
"""

from openpyxl import load_workbook, Workbook


class JobRanker:

    def __init__(self):

        self.company_file = "data/companies.xlsx"
        self.database_file = "data/jobs_database.xlsx"
        self.output_file = "data/todays_top_jobs.xlsx"

        self.keywords = self.load_keywords()

    def load_keywords(self):

        wb = load_workbook(self.company_file)

        ws = wb["Keywords"]

        keywords = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            if row[0]:
                keywords.append(str(row[0]).lower())

        return keywords

    def score_job(self, title, description):

        score = 0

        matched = []

        text = f"{title} {description}".lower()

        for keyword in self.keywords:

            if keyword in text:

                score += 5

                matched.append(keyword)

        if score > 100:
            score = 100

        return score, ", ".join(matched)

    def rank_jobs(self):

        db = load_workbook(self.database_file)

        ws = db.active

        out = Workbook()

        result = out.active

        result.title = "Today's Top Jobs"

        result.append([

            "Score",

            "Company",

            "Job Title",

            "Country",

            "Status",

            "Matched Keywords",

            "Apply URL"

        ])

        jobs = []

        for row in ws.iter_rows(min_row=2, values_only=True):

            company = row[1]
            title = row[2]
            country = row[3]
            url = row[6]
            description = row[7]
            status = row[11]

            score, matched = self.score_job(title, description)

            if score == 0:
                continue

            jobs.append([

                score,
                company,
                title,
                country,
                status,
                matched,
                url

            ])

        jobs.sort(reverse=True, key=lambda x: x[0])

        for job in jobs:

            result.append(job)

        out.save(self.output_file)

        print(f"\nTop Jobs Saved : {len(jobs)}")